#!/usr/bin/env python3
"""
Módulo 3 — Cotizador Automático
Marc Chagall Laboratorio Fotográfico — Salta, Argentina

Para cada ítem de una licitación:
  1. Busca precio de referencia en MercadoLibre y/o PrecioClaro
  2. Calcula precio de venta con margen configurable
  3. Genera Excel de cotización con openpyxl
"""

import argparse
import json
import logging
import re
import time
import urllib.parse
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("logs/cotizador.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

BASE_DIR = Path(__file__).parent
LICITACIONES_FILE = BASE_DIR / "licitaciones.json"
COTIZACIONES_DIR = BASE_DIR / "cotizaciones"
CONFIG_FILE = BASE_DIR / "config.json"

COTIZACIONES_DIR.mkdir(exist_ok=True)
(BASE_DIR / "logs").mkdir(exist_ok=True)

# ── Config ────────────────────────────────────────────────────────────────────

def load_config() -> dict:
    with open(CONFIG_FILE, encoding="utf-8") as f:
        return json.load(f)


def load_licitaciones() -> list[dict]:
    with open(LICITACIONES_FILE, encoding="utf-8") as f:
        return json.load(f)


def save_licitaciones(data: list[dict]) -> None:
    with open(LICITACIONES_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── Headers HTTP ──────────────────────────────────────────────────────────────
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "es-AR,es;q=0.9",
}
SESSION = requests.Session()
SESSION.headers.update(HEADERS)


def get(url: str, timeout: int = 15) -> requests.Response | None:
    try:
        r = SESSION.get(url, timeout=timeout)
        r.raise_for_status()
        time.sleep(1.0)
        return r
    except requests.RequestException as e:
        log.warning("GET %s → %s", url, e)
        return None


# ── Búsqueda de precios ───────────────────────────────────────────────────────

def buscar_mercadolibre(descripcion: str) -> tuple[float | None, str | None]:
    """
    Busca el precio más barato de un ítem en MercadoLibre Argentina.
    Devuelve (precio_sin_iva, url_producto).
    """
    query = urllib.parse.quote(descripcion)
    url = f"https://listado.mercadolibre.com.ar/{query}"
    log.info("  ML: buscando '%s'", descripcion[:60])

    resp = get(url)
    if not resp:
        return None, None

    soup = BeautifulSoup(resp.text, "lxml")

    # Selectores de MercadoLibre Argentina (pueden cambiar con rediseños)
    precios = []
    selectors = [
        "span.andes-money-amount__fraction",
        "span.price-tag-fraction",
        "[class*='price__fraction']",
        "[class*='money-amount__fraction']",
    ]
    for sel in selectors:
        tags = soup.select(sel)
        if tags:
            for tag in tags[:10]:
                raw = tag.get_text(strip=True).replace(".", "").replace(",", "")
                try:
                    precios.append(float(raw))
                except ValueError:
                    pass
            break

    if not precios:
        log.warning("  ML: sin precios para '%s'", descripcion[:60])
        return None, None

    precio_min = min(precios)
    # ML muestra precios con IVA incluido para consumidor final
    # Para precio sin IVA dividimos por 1.21
    precio_sin_iva = round(precio_min / 1.21, 2)
    log.info("  ML: precio mín encontrado $%.2f (con IVA) → $%.2f (sin IVA)", precio_min, precio_sin_iva)
    return precio_sin_iva, url


def buscar_precio_claro(descripcion: str) -> tuple[float | None, str | None]:
    """
    Fallback: busca en PrecioClaro del gobierno argentino.
    Devuelve (precio_sin_iva, url).
    """
    query = urllib.parse.quote(descripcion)
    url = f"https://www.precioclaro.gob.ar/precioclaro/rest/precioClaro?cadena=0&query={query}&limit=10"
    log.info("  PrecioClaro: buscando '%s'", descripcion[:60])

    resp = get(url)
    if not resp:
        return None, None

    try:
        data = resp.json()
        productos = data.get("data", {}).get("productos", []) or data.get("productos", [])
        if not productos:
            return None, None

        precios_validos = []
        for prod in productos:
            precio = prod.get("precio") or prod.get("precioMin")
            if precio:
                try:
                    precios_validos.append(float(str(precio).replace(",", ".")))
                except ValueError:
                    pass

        if not precios_validos:
            return None, None

        precio_min = min(precios_validos)
        url_fuente = "https://www.precioclaro.gob.ar"
        log.info("  PrecioClaro: precio mín $%.2f", precio_min)
        return round(precio_min, 2), url_fuente
    except Exception as e:
        log.warning("  PrecioClaro: error parseando respuesta: %s", e)
        return None, None


def buscar_precio_google(descripcion: str, api_key: str, cse_id: str) -> tuple[float | None, str | None]:
    """
    Busca precio usando Google Custom Search API (opcional).
    """
    if not api_key or not cse_id:
        return None, None
    query = f"{descripcion} precio Argentina"
    url = (
        f"https://www.googleapis.com/customsearch/v1"
        f"?key={api_key}&cx={cse_id}&q={urllib.parse.quote(query)}&num=5"
    )
    resp = get(url)
    if not resp:
        return None, None
    try:
        items = resp.json().get("items", [])
        for item in items:
            snippet = item.get("snippet", "")
            m = re.search(r"\$\s?([\d.,]+)", snippet)
            if m:
                precio = float(m.group(1).replace(".", "").replace(",", "."))
                return round(precio / 1.21, 2), item.get("link")
    except Exception as e:
        log.warning("Google CSE error: %s", e)
    return None, None


def obtener_precio_referencia(descripcion: str, config: dict) -> tuple[float | None, str | None]:
    """
    Orquesta la búsqueda de precios en orden de prioridad.
    Devuelve (precio_sin_iva, fuente).
    """
    # 1. MercadoLibre
    precio, fuente = buscar_mercadolibre(descripcion)
    if precio:
        return precio, f"MercadoLibre: {fuente}"

    # 2. PrecioClaro
    precio, fuente = buscar_precio_claro(descripcion)
    if precio:
        return precio, f"PrecioClaro: {fuente}"

    # 3. Google CSE (opcional)
    precio, fuente = buscar_precio_google(
        descripcion,
        config.get("google_api_key", ""),
        config.get("google_cse_id", ""),
    )
    if precio:
        return precio, f"Google: {fuente}"

    return None, None


# ── Generación del Excel ──────────────────────────────────────────────────────

# Paleta Marc Chagall
COLOR_FONDO      = "0A0C10"
COLOR_DORADO     = "C8A84B"
COLOR_DORADO_CLR = "C8A84B"
COLOR_HEADER_BG  = "1A1D24"
COLOR_FILA_PAR   = "13161E"
COLOR_FILA_IMP   = "0F1118"
COLOR_TEXTO      = "E8E8E8"
COLOR_VERDE      = "2ECC71"
COLOR_ROJO       = "E74C3C"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=11, color=COLOR_TEXTO, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)


def _border_thin() -> Border:
    lado = Side(style="thin", color="2A2D35")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def generar_excel(licitacion: dict, items_cotizados: list[dict], config: dict) -> Path:
    """Genera el Excel de cotización y devuelve la ruta."""
    wb = Workbook()

    # ── Hoja 1: Cotización ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Cotización"
    ws.sheet_view.showGridLines = False
    ws.tab_color = COLOR_DORADO

    # Anchos de columna
    anchos = [6, 50, 12, 12, 16, 16, 10, 18, 18, 25]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    fila = 1

    # ── Encabezado empresa ────────────────────────────────────────────────────
    ws.row_dimensions[fila].height = 40
    ws.merge_cells(f"A{fila}:J{fila}")
    c = ws.cell(fila, 1, config.get("empresa", "Marc Chagall Laboratorio Fotográfico").upper())
    c.fill = _fill(COLOR_FONDO)
    c.font = Font(name="Calibri", bold=True, size=18, color=COLOR_DORADO)
    c.alignment = _center()
    fila += 1

    ws.row_dimensions[fila].height = 20
    ws.merge_cells(f"A{fila}:J{fila}")
    c = ws.cell(fila, 1, f"COTIZACIÓN DE LICITACIÓN")
    c.fill = _fill(COLOR_FONDO)
    c.font = Font(name="Calibri", bold=True, size=12, color="888888")
    c.alignment = _center()
    fila += 1

    # ── Info licitación ────────────────────────────────────────────────────────
    info_rows = [
        ("Organismo:", licitacion.get("organismo", "")),
        ("Título:", licitacion.get("titulo", "")),
        ("Expediente:", licitacion.get("expediente", "")),
        ("Fecha Apertura:", licitacion.get("fecha_apertura", "")[:16] if licitacion.get("fecha_apertura") else ""),
        ("Fecha Cotización:", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    for label, valor in info_rows:
        ws.row_dimensions[fila].height = 18
        ws.merge_cells(f"A{fila}:C{fila}")
        c = ws.cell(fila, 1, label)
        c.fill = _fill(COLOR_HEADER_BG)
        c.font = _font(bold=True, color=COLOR_DORADO, size=10)
        c.alignment = _left()

        ws.merge_cells(f"D{fila}:J{fila}")
        c = ws.cell(fila, 4, valor)
        c.fill = _fill(COLOR_HEADER_BG)
        c.font = _font(size=10)
        c.alignment = _left()
        fila += 1

    fila += 1  # espacio

    # ── Header tabla ──────────────────────────────────────────────────────────
    headers = [
        "N°", "Descripción", "Cant.", "Unidad",
        "Precio Ref\n(sin IVA)", "Precio Ref\n(con IVA)",
        "Margen\n%", "P. Oferta\nUnit.", "P. Oferta\nTotal", "Fuente Precio"
    ]
    ws.row_dimensions[fila].height = 36
    for col, header in enumerate(headers, 1):
        c = ws.cell(fila, col, header)
        c.fill = _fill(COLOR_DORADO)
        c.font = Font(name="Calibri", bold=True, size=10, color=COLOR_FONDO)
        c.alignment = _center()
        c.border = _border_thin()
    header_fila = fila
    fila += 1

    # ── Filas de ítems ─────────────────────────────────────────────────────────
    iva = config.get("iva", 21) / 100
    margen_pct = items_cotizados[0].get("margen_pct", 25) if items_cotizados else 25

    fila_items_inicio = fila
    for idx, item in enumerate(items_cotizados):
        ws.row_dimensions[fila].height = 20
        color_bg = COLOR_FILA_PAR if idx % 2 == 0 else COLOR_FILA_IMP

        precio_ref_sin_iva = item.get("precio_referencia")
        precio_ref_con_iva = round(precio_ref_sin_iva * (1 + iva), 2) if precio_ref_sin_iva else None
        margen = item.get("margen_pct", margen_pct)
        cantidad = item.get("cantidad") or 1

        if precio_ref_sin_iva:
            precio_oferta_unit = round(precio_ref_sin_iva * (1 + margen / 100), 2)
            precio_oferta_total = round(precio_oferta_unit * cantidad, 2)
        else:
            precio_oferta_unit = None
            precio_oferta_total = None

        valores = [
            item.get("numero", idx + 1),
            item.get("descripcion", ""),
            cantidad,
            item.get("unidad", ""),
            precio_ref_sin_iva,
            precio_ref_con_iva,
            f"{margen}%",
            precio_oferta_unit,
            precio_oferta_total,
            item.get("fuente_precio", ""),
        ]

        for col, val in enumerate(valores, 1):
            c = ws.cell(fila, col, val)
            c.fill = _fill(color_bg)
            c.font = _font(size=10)
            c.border = _border_thin()
            if col in (5, 6, 8, 9) and val is not None:
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif col == 3:
                c.alignment = _center()
                c.number_format = '#,##0.##'
            elif col == 2:
                c.alignment = _left()
            else:
                c.alignment = _center()

            # Resaltar sin precio en rojo suave
            if col == 8 and val is None:
                c.font = Font(name="Calibri", size=10, color="888888", italic=True)
                c.value = "Sin cotizar"

        fila += 1

    fila_items_fin = fila - 1
    fila += 1  # espacio

    # ── Resumen financiero ────────────────────────────────────────────────────
    resumen_inicio = fila

    # Calcular totales
    total_sin_iva = sum(
        (it.get("precio_referencia") or 0) * (it.get("cantidad") or 1)
        * (1 + (it.get("margen_pct", margen_pct)) / 100)
        for it in items_cotizados
        if it.get("precio_referencia")
    )
    total_iva = round(total_sin_iva * iva, 2)
    total_con_iva = round(total_sin_iva + total_iva, 2)
    ganancia = round(total_sin_iva * (margen_pct / (100 + margen_pct)), 2)
    costo_total = total_sin_iva - ganancia

    resumen_data = [
        ("SUBTOTAL SIN IVA", total_sin_iva, False),
        (f"IVA ({config.get('iva', 21)}%)", total_iva, False),
        ("TOTAL CON IVA", total_con_iva, True),
        ("", "", False),
        ("Costo estimado (sin ganancia)", costo_total, False),
        (f"Ganancia estimada ({margen_pct}%)", ganancia, False),
        ("Ganancia / Ingresos", f"{round(ganancia / total_sin_iva * 100, 1)}%" if total_sin_iva else "—", False),
    ]

    for label, valor, destacado in resumen_data:
        if label == "":
            fila += 1
            continue
        ws.row_dimensions[fila].height = 22
        ws.merge_cells(f"F{fila}:H{fila}")
        c = ws.cell(fila, 6, label)
        c.fill = _fill(COLOR_DORADO if destacado else COLOR_HEADER_BG)
        c.font = Font(
            name="Calibri", bold=destacado, size=11,
            color=COLOR_FONDO if destacado else (COLOR_DORADO if not destacado else COLOR_TEXTO)
        )
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = _border_thin()

        c2 = ws.cell(fila, 9, valor)
        c2.fill = _fill(COLOR_DORADO if destacado else COLOR_HEADER_BG)
        c2.font = Font(
            name="Calibri", bold=destacado, size=11,
            color=COLOR_FONDO if destacado else COLOR_VERDE
        )
        c2.alignment = Alignment(horizontal="right", vertical="center")
        c2.border = _border_thin()
        if isinstance(valor, float):
            c2.number_format = '$ #,##0.00'
        fila += 1

    # ── Hoja 2: Detalle de precios ────────────────────────────────────────────
    ws2 = wb.create_sheet("Fuentes de Precios")
    ws2.sheet_view.showGridLines = False
    ws2.tab_color = "888888"

    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 60

    ws2.row_dimensions[1].height = 30
    ws2.merge_cells("A1:D1")
    c = ws2.cell(1, 1, "DETALLE DE FUENTES DE PRECIOS")
    c.fill = _fill(COLOR_FONDO)
    c.font = Font(name="Calibri", bold=True, size=14, color=COLOR_DORADO)
    c.alignment = _center()

    for col, h in enumerate(["N°", "Descripción", "Precio Ref (sin IVA)", "Fuente / URL"], 1):
        c = ws2.cell(2, col, h)
        c.fill = _fill(COLOR_DORADO)
        c.font = Font(name="Calibri", bold=True, size=10, color=COLOR_FONDO)
        c.alignment = _center()
        c.border = _border_thin()

    for i, item in enumerate(items_cotizados, 1):
        for col, val in enumerate([
            i,
            item.get("descripcion", ""),
            item.get("precio_referencia"),
            item.get("fuente_precio", ""),
        ], 1):
            c = ws2.cell(i + 2, col, val)
            c.fill = _fill(COLOR_FILA_PAR if i % 2 == 0 else COLOR_FILA_IMP)
            c.font = _font(size=9)
            c.border = _border_thin()
            if col == 3 and val:
                c.number_format = '$ #,##0.00'
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = _left()

    # ── Guardar ───────────────────────────────────────────────────────────────
    nombre_lic = re.sub(r"[^\w\s\-]", "", licitacion.get("titulo", licitacion["id"]))[:50]
    nombre_lic = re.sub(r"\s+", "-", nombre_lic.strip())
    nombre_archivo = f"{nombre_lic}-cotizacion.xlsx"
    ruta = COTIZACIONES_DIR / nombre_archivo
    wb.save(ruta)
    log.info("Excel generado: %s", ruta)
    return ruta


# ── Main ──────────────────────────────────────────────────────────────────────

def cotizar_licitacion(lic_id: str, margen: float, force: bool = False) -> Path | None:
    config = load_config()
    licitaciones = load_licitaciones()

    lic = next((l for l in licitaciones if l["id"] == lic_id), None)
    if not lic:
        log.error("Licitación '%s' no encontrada en licitaciones.json", lic_id)
        return None

    items = lic.get("items", [])
    if not items:
        log.error("[%s] No tiene ítems. Ejecutá primero: python extraer_items.py", lic_id)
        return None

    log.info("[%s] Cotizando %d ítems con margen %.1f%%", lic_id, len(items), margen)

    items_cotizados = []
    for item in items:
        if not force and item.get("precio_referencia"):
            log.info("  [ya cotizado] %s", item["descripcion"][:50])
            items_cotizados.append({**item, "margen_pct": margen})
            continue

        desc = item.get("descripcion", "")
        precio, fuente = obtener_precio_referencia(desc, config)
        items_cotizados.append({
            **item,
            "precio_referencia": precio,
            "fuente_precio": fuente,
            "margen_pct": margen,
        })

    # Guardar precios en licitaciones.json
    lic["items"] = [
        {**orig, "precio_referencia": cot["precio_referencia"], "fuente_precio": cot["fuente_precio"]}
        for orig, cot in zip(items, items_cotizados)
    ]
    lic["cotizacion"] = {
        "fecha": datetime.now().isoformat(),
        "margen": margen,
        "items_cotizados": len([i for i in items_cotizados if i.get("precio_referencia")]),
        "items_sin_precio": len([i for i in items_cotizados if not i.get("precio_referencia")]),
    }

    # Guardar en JSON
    for i, l in enumerate(licitaciones):
        if l["id"] == lic_id:
            licitaciones[i] = lic
            break
    save_licitaciones(licitaciones)

    # Generar Excel
    ruta_excel = generar_excel(lic, items_cotizados, config)
    lic["cotizacion"]["excel"] = str(ruta_excel.relative_to(BASE_DIR))
    for i, l in enumerate(licitaciones):
        if l["id"] == lic_id:
            licitaciones[i] = lic
            break
    save_licitaciones(licitaciones)

    print(f"\n✅ Cotización completada para '{lic.get('titulo', lic_id)}'")
    print(f"   Excel: {ruta_excel}")
    sin_precio = [i for i in items_cotizados if not i.get("precio_referencia")]
    if sin_precio:
        print(f"   ⚠️  {len(sin_precio)} ítems sin precio (requieren cotización manual):")
        for sp in sin_precio:
            print(f"      - {sp['descripcion'][:70]}")

    return ruta_excel


def main():
    parser = argparse.ArgumentParser(
        description="Cotizador automático de licitaciones — Marc Chagall"
    )
    parser.add_argument("--id", required=True, help="ID de la licitación")
    parser.add_argument("--margen", type=float, default=None, help="Margen %% (default: config.json)")
    parser.add_argument("--force", action="store_true", help="Forzar re-cotización aunque ya tenga precios")
    args = parser.parse_args()

    config = load_config()
    margen = args.margen if args.margen is not None else config.get("margen_default", 25)

    log.info("=" * 60)
    log.info("Cotizador — %s", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    log.info("=" * 60)

    cotizar_licitacion(args.id, margen, force=args.force)


if __name__ == "__main__":
    main()
